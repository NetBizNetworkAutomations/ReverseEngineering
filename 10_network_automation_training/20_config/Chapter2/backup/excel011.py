from openpyxl import load_workbook

filename = 'sample.xlsx'
wb = load_workbook(filename)
ws = wb['Shizuoka']

data = ws.cell(6, 4).value

print(data)