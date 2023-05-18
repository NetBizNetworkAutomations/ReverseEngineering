from openpyxl import load_workbook

wb = load_workbook('sample.xlsx')
wb.create_sheet('test_sheet')

wb.save('sample.xlsx')