from openpyxl import load_workbook

filename = 'sample.xlsx'
wb = load_workbook(filename)
ws = wb['Tokyo']

ws.cell(7, 3).value = '172.16.100.100'

wb.save(filename)