from openpyxl import load_workbook

filename = 'sample.xlsx'
wb = load_workbook(filename)
ws = wb['Tokyo']

ws.cell(row = 7, column = 3).value = '10.0.0.1'

wb.save(filename)