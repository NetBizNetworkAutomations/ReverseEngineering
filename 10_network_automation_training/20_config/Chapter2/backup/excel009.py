from openpyxl import load_workbook

filename = 'sample.xlsx'
wb = load_workbook(filename)
ws = wb['Shizuoka']

for x in range(12, 16):
    ws.cell(x, 5).value = 'shutdown'

wb.save(filename)