from openpyxl import load_workbook

filename = 'sample.xlsx'
wb = load_workbook(filename)
ws = wb['Tokyo']

ws['C3'] = 'New_Tokyo_Router'

wb.save(filename)