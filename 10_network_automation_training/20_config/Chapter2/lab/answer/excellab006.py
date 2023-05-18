from openpyxl import load_workbook

HOSTNAME = 'C3'

filename = 'switch.xlsx'
wb = load_workbook(filename)

sheets = wb.sheetnames

for sheet in sheets:
    ws = wb[sheet]
    ws[HOSTNAME] = sheet

wb.save(filename)