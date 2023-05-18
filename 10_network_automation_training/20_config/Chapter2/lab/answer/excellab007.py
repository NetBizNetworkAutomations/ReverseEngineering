from openpyxl import load_workbook
import re

IP = 'C42'

filename = 'switch.xlsx'
wb = load_workbook(filename)

sheets = wb.sheetnames

for sheet in sheets:
    ws = wb[sheet]
    m = re.search('(_)(\d+)', sheet)
    ws[IP] = '192.168.200.' + m.group(2)

wb.save(filename)