from openpyxl import load_workbook

filename = 'switch.xlsx'
wb = load_workbook(filename)
ws = wb['Tokyo_3F_SW']

print('変更前のSTPモード：' + ws['C4'].value)
print('変更前のVTPモード：' + ws['D5'].value)
print('変更前のVTPドメイン名：' + ws['D6'].value)
print('変更前のイネーブルシークレット' + ws['D8'].value)

print('-'*40)

ws['C4'] = 'rapid-pvst'
ws['D5'] = 'server'
ws['D6'] = 'abc123'
ws['D8'] = 'san-fran'

print('変更後のSTPモード：' + ws['C4'].value)
print('変更後のVTPモード：' + ws['D5'].value)
print('変更後のVTPドメイン名：' + ws['D6'].value)
print('変更後のイネーブルシークレット：' + ws['D8'].value)

wb.save(filename)