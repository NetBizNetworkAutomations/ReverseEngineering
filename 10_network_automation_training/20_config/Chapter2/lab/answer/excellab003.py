from openpyxl import load_workbook

filename = 'switch.xlsx'
wb = load_workbook(filename)
ws = wb['Tokyo_2F_SW']

for x in range(15, 38):
    ws.cell(x, 4).value = 50

for x in range(15, 39):
    ws.cell(x, 5).value = 'shutdown'

wb.save(filename)