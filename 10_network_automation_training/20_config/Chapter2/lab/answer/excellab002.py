from openpyxl import load_workbook

filename = 'switch.xlsx'
wb = load_workbook(filename)
ws = wb['Tokyo_7F_SW']

print('変更前のSTPモード：' + ws.cell(4, 3).value)
print('変更前のVTPモード：' + ws.cell(5, 4).value)
print('変更前のVTPドメイン名：' + ws.cell(6, 4).value)
print('変更前のイネーブルシークレット' + ws.cell(8, 4).value)

print('-'*40)

ws.cell(4, 3).value = 'rapid-pvst'
ws.cell(5, 4).value = 'server'
ws.cell(6, 4).value = 'abc123'
ws.cell(8, 4).value = 'san-fran'

print('変更後のSTPモード：' + ws.cell(4, 3).value)
print('変更後のVTPモード：' + ws.cell(5, 4).value)
print('変更後のVTPドメイン名：' + ws.cell(6, 4).value)
print('変更後のイネーブルシークレット：' + ws.cell(8, 4).value)

wb.save(filename)