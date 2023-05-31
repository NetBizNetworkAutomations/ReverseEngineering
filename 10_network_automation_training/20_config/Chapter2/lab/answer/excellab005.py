from openpyxl import load_workbook

NTP = 'C10'

filename = 'switch.xlsx'
wb = load_workbook(filename)

sheets = wb.sheetnames

for sheet in sheets:
    ws = wb[sheet]
    ws[NTP] = '172.16.1.254'

wb.save(filename)