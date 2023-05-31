from openpyxl import load_workbook

filename = 'switch.xlsx'
wb = load_workbook(filename)
ws = wb['Tokyo_3F_SW']

for x in range(15, 38):
    if x < 20:
        ws.cell(x, 4).value = 200
    elif x < 27:
        ws.cell(x, 4).value = 300
    else:
        ws.cell(x, 4).value = 400

wb.save(filename)