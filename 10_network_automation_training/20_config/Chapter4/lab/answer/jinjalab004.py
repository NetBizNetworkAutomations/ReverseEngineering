from jinja2 import Environment, FileSystemLoader
from openpyxl import load_workbook

wb = load_workbook('ctct_network.xlsx')
ws = wb['komazawa']

data = {
    'name': ws['C3'].value,
    'secret_pass': ws['D5'].value,
    'int_name': ws['B13'].value,
    'ip_address': ws['C13'].value,
    'mask': ws['D13'].value,
    'status': ws['E13'].value,
    'console_pass': ws['D4'].value,
    'route': [],
}

for x in range(21, 26):
    if ws.cell(x, 2).value:
        data['route'].append({
            'ip': ws.cell(x, 2).value,
            'mask': ws.cell(x, 3).value,
            'next_hop': ws.cell(x, 4).value,
        })
    else:
        break

env = Environment(loader = FileSystemLoader('./template'))
tmpl = env.get_template('komazawa.j2')

config = tmpl.render(data)

with open('komazawa_config.cfg', 'w') as f:
    f.write(config)