from jinja2 import Environment, FileSystemLoader
from openpyxl import load_workbook

wb = load_workbook('sample3.xlsx')
sheets = wb.sheetnames

for sheet in sheets:
    ws = wb[sheet]
    data = {
        'name': ws['C3'].value,
        'int': [],
    }
    
    for x in range(12, 16):
        if ws.cell(x, 2):
            data['int'].append({
                'name': ws.cell(x, 2).value,
                'ip_address': ws.cell(x, 3).value,
                'mask': ws.cell(x, 4).value,
            })
        else:
            break
    
    env = Environment(loader = FileSystemLoader('./template'))
    tmpl = env.get_template('int_config.j2')
    
    config = tmpl.render(data)
    
    with open(data['name'] + '.cfg', 'w') as f:
        f.write(config)