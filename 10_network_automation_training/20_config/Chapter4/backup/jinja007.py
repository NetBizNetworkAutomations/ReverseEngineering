from openpyxl import load_workbook

wb = load_workbook('sample3.xlsx')
sheets = wb.sheetnames

for sheet in sheets:
    ws = wb[sheet]
    print(ws['C3'].value)
