from jinja2 import Environment, FileSystemLoader
from openpyxl import load_workbook

wb = load_workbook('sample3.xlsx')
ws = wb['Tokyo']

data = {
    'host': ws['C3'].value,
    'console_pass': ws['D4'].value,
    'secret_pass': ws['D5'].value,
    'ip_address': ws['C13'].value,
    'mask': ws['D13'].value,
}

env = Environment(loader = FileSystemLoader('./template'))
tmpl = env.get_template('base_config.j2')

config = tmpl.render(data)

print(config)