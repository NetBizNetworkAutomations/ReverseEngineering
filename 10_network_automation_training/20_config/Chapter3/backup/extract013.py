from netaddr import IPNetwork
from openpyxl import load_workbook

filename = 'sample2.xlsx'
wb = load_workbook(filename)
ws = wb['Tokyo']

address = IPNetwork('192.168.100.1/24')

ws.cell(17, 2).value = 'GigabitEthernet0/0'
ws.cell(17, 3).value = 'no shutdown'
ws.cell(17, 4).value = str(address.ip)
ws.cell(17, 5).value = str(address.netmask)

wb.save(filename)