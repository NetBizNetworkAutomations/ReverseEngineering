from netmiko import ConnectHandler
from openpyxl import load_workbook

device = {
    'device_type': 'cisco_ios_telnet',
    'host': '10.1.1.254',
    'password': 'python',
    'secret': 'cisco',
}

net_connect = ConnectHandler(**device)
net_connect.enable()

cmd = 'show interfaces'
result = net_connect.send_command(cmd, use_textfsm = True)

net_connect.disconnect()

filename = 'sample2.xlsx'
wb = load_workbook(filename)
ws = wb['Tokyo']

row =  17
for x in result:
    emp = '-'
    
    ws.cell(row, 2).value = x['interface']
    ws.cell(row, 3).value = x['link_status']
    ws.cell(row, 4).value = x['ip_address'] or emp
    
    row += 1

wb.save(filename)