from netmiko import ConnectHandler
from openpyxl import load_workbook

device = {
    'device_type': 'cisco_ios_telnet',
    'host': '10.1.1.254',
    'password': 'python',
    'secret': 'cisco',
}

net_connect = ConnectHandler(**device)

net_connect.enable()
cmd = 'show running-config'
result = net_connect.send_command(cmd)

net_connect.disconnect()

contents = result.splitlines()

for content in contents:
    if 'timezone' in content:
        timezone = content.split()[2]

filename = 'sample2.xlsx'
wb = load_workbook(filename)
ws = wb['Tokyo']

ws['C13'] = timezone

wb.save(filename)