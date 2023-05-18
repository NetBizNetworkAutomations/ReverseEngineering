from netmiko import ConnectHandler
from openpyxl import load_workbook

device = {
    'device_type': 'cisco_ios_telnet',
    'host': '10.1.1.254',
    'password': 'python',
    'secret': 'cisco',
}

net_connect = ConnectHandler(**device)

net_connect.enable()

cmd = 'show version'
result = net_connect.send_command(cmd, use_textfsm = True)

for x in result:
    hardware = x['hardware'][0]
    register = x['config_register']

net_connect.disconnect()

filename = 'sample2.xlsx'
wb = load_workbook(filename)
ws = wb['Tokyo']

ws['C5'] = hardware
ws['C6'] = register

wb.save(filename)