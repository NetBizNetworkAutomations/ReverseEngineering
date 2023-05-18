from netmiko import ConnectHandler
from openpyxl import load_workbook
import re

device = {
    'device_type': 'cisco_ios_telnet',
    'host': '10.1.1.254',
    'password': 'python',
    'secret': 'cisco',
}

net_connect = ConnectHandler(**device)

net_connect.enable()
cmd = 'show running-config'
result = net_connect.send_command(cmd)

net_connect.disconnect()

contents = result.splitlines()

mode = False
for content in contents:
    if 'line con' in content:
        mode = True
    elif mode and re.match(' ', content):
        if 'password' in content:
            console_pass = content.split()[1]

filename = 'sample2.xlsx'
wb = load_workbook(filename)
ws = wb['Tokyo']

ws['D10'] = console_pass

wb.save(filename)