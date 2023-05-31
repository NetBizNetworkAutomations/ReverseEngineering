from netmiko import ConnectHandler
from openpyxl import load_workbook

device = {
    'device_type': 'cisco_ios_telnet',
    'host': '10.1.1.254',
    'password': 'python',
    'secret': 'cisco',
}

net_connect = ConnectHandler(**device)

net_connect.enable()

cmd = 'show running-config'
result = net_connect.send_command(cmd)

net_connect.disconnect()

contents = result.splitlines()

route = []
for content in contents:
    if 'ip route' in content:
        route.append(content)

filename = 'Tokyo_Router.xlsx'
wb = load_workbook(filename)
ws = wb['sheet1']

row = 36
for x in route:
    ws.cell(row, 2).value = x.split()[2]
    ws.cell(row, 3).value = x.split()[3]
    ws.cell(row, 4).value = x.split()[4]
    row += 1

wb.save(filename)