from netmiko import ConnectHandler
from openpyxl import load_workbook

device = {
    'device_type': 'cisco_ios_telnet',
    'host': '10.1.1.254',
    'password': 'python',
    'secret': 'cisco',
}

net_connect = ConnectHandler(**device)

net_connect.enable()

cmd = 'show version'
result = net_connect.send_command(cmd, use_textfsm = True)

net_connect.disconnect()

filename = 'Tokyo_Router.xlsx'
wb = load_workbook(filename)
ws = wb['sheet1']

for x in result:
    ws['C3'] = x['hardware'][0]
    ws['C4'] = x['serial'][0]
    ws['C5'] = x['version']
    ws['C6'] = x['config_register']

wb.save(filename)