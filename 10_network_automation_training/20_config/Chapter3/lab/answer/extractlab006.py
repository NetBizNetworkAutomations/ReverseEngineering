from netmiko import ConnectHandler
from openpyxl import load_workbook

device = {
    'device_type': 'cisco_ios_telnet',
    'host': '10.1.1.254',
    'password': 'python',
    'secret': 'cisco',
}

net_connect = ConnectHandler(**device)

net_connect.enable()

cmd = 'show inventory'
result = net_connect.send_command(cmd, use_textfsm = True)

net_connect.disconnect()

filename = 'Tokyo_Router.xlsx'
wb = load_workbook(filename)
ws = wb['sheet1']

row = 16
for x in result:
    ws.cell(row, 2).value = x['pid']
    ws.cell(row, 3).value = x['sn']
    ws.cell(row, 4).value = x['descr']
    row += 1

wb.save(filename)