from netmiko import ConnectHandler
from openpyxl import load_workbook
from netaddr import IPNetwork

device = {
    'device_type': 'cisco_ios_telnet',
    'host': '10.1.1.254',
    'password': 'python',
    'secret': 'cisco',
}

net_connect = ConnectHandler(**device)

net_connect.enable()

cmd = 'show interfaces'
result = net_connect.send_command(cmd, use_textfsm = True)

net_connect.disconnect()

filename = 'Tokyo_Router.xlsx'
wb = load_workbook(filename)
ws = wb['sheet1']

row = 24
for x in result:
    ws.cell(row, 2).value = x['interface']
    ws.cell(row, 3).value = x['link_status']
    
    if x['ip_address']:
        address = IPNetwork(x['ip_address'])
        ws.cell(row, 4).value = str(address.ip)
        ws.cell(row, 5).value = str(address.netmask)
    else:
        ws.cell(row, 4).value = '-'
        ws.cell(row, 5).value = '-'
    
    row += 1

wb.save(filename)