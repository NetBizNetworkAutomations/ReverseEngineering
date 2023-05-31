from netmiko import ConnectHandler
from openpyxl import load_workbook
from netaddr import IPNetwork

device = {
    'device_type': 'cisco_ios_telnet',
    'host': '10.1.1.254',
    'password': 'python',
    'secret': 'cisco',
}

net_connect = ConnectHandler(**device)

net_connect.enable()

cmd = 'show ip route'
result = net_connect.send_command(cmd, use_textfsm = True)

net_connect.disconnect()

filename = 'Tokyo_Router.xlsx'
wb = load_workbook(filename)
ws = wb['sheet1']

row = 49
for x in result:
    ws.cell(row, 2).value = x['protocol']
    ws.cell(row, 3).value = x['network']
    
    mask = IPNetwork('0.0.0.0/' + x['mask'])
    
    ws.cell(row, 4).value = str(mask.netmask)
    ws.cell(row, 5).value = x['nexthop_ip'] or '-'
    ws.cell(row, 6).value = x['nexthop_if'] or '-'
    
    row += 1

wb.save(filename)