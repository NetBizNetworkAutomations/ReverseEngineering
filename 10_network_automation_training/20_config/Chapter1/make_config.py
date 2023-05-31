from jinja2 import Environment, FileSystemLoader
from openpyxl import load_workbook

wb = load_workbook('education_network.xlsx')
sheets = wb.sheetnames

for sheet in sheets:
    ws = wb[sheet]
    
    data = {
        'name': ws['C3'].value,
        'secret_pass': ws['D5'].value,
        'process': ws['C28'].value,
        'router_id': ws['C29'].value,
        'console_pass': ws['D4'].value,
        'vty_pass': ws['D6'].value,
        'int': [],
        'ospf': [],
    }
    
    for x in range(12, 17):
        if ws.cell(x, 2):
            data['int'].append({
                'int_name': ws.cell(x, 2).value,
                'ip_address': ws.cell(x, 3).value,
                'mask': ws.cell(x, 4).value,
                'status': ws.cell(x, 5).value,
                'desc': ws.cell(x, 6).value,
            })
        else:
            break
    
    for x in range(31, 35):
        if ws.cell(x, 2):
            data['ospf'].append({
                'network': ws.cell(x, 2).value,
                'wildmask': ws.cell(x, 3).value,
                'area': ws.cell(x, 4).value,
            })
        else:
            break
    
    filename = ws.title + '.cfg'
    with open(filename, 'w') as f:
        env = Environment(loader = FileSystemLoader('.'))
        tmpl = env.get_template('education_network.j2')
        
        config = tmpl.render(data)
        
        f.write(config)