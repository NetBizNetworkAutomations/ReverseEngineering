from netmiko import ConnectHandler
from openpyxl import load_workbook
from netaddr import IPNetwork
import re

R1 = {
    'device_type': 'cisco_ios_telnet',
    'host': '10.1.1.254',
    'password': 'python',
    'secret': 'cisco',
}

print('R1への接続を開始します')
net_connect = ConnectHandler(**R1)
net_connect.enable()

# 「基本情報」の取得
print('「基本情報」を取得しています')
cmd = 'show version'
basic_data = net_connect.send_command(cmd, use_textfsm = True)

# 「設定情報」と「スタティックルート設定情報」の取得
print('「設定情報」と「スタティックルート設定情報」を取得しています')
cmd = 'show running-config'
result = net_connect.send_command(cmd)
config_static_data = result.splitlines()

# 「モジュール情報」の取得
print('「モジュール情報」を取得しています')
cmd = 'show inventory'
module_data = net_connect.send_command(cmd, use_textfsm = True)

# 「インタフェース情報」の取得
print('「インタフェース情報」を取得しています')
cmd = 'show interfaces'
interface_data = net_connect.send_command(cmd, use_textfsm = True)

# 「ルーティングテーブル情報」の取得
print('「ルーティングテーブル情報」を取得しています')
cmd = 'show ip route'
route_data = net_connect.send_command(cmd, use_textfsm = True)

net_connect.disconnect()
print('R1から切断しました')

# R1.xlsxへの書き込み処理
print('取得した情報のエクセルへの書き込みを開始します')
filename = 'R1.xlsx'
wb = load_workbook(filename)
ws = wb['sheet1']

# 「基本情報」の書き込み
for x in basic_data:
    ws['C3'] = x['hardware'][0]
    ws['C4'] = x['serial'][0]
    ws['C5'] = x['version']
    ws['C6'] = x['config_register']

# 「設定情報」の書き込み
mode = False
for x in config_static_data:
    if 'hostname' in x:
        ws['C9'] = x.split()[1]
    elif 'ntp server' in x:
        ws['C11'] = x.split()[2]
    elif 'clock timezone' in x:
        ws['C12'] = x.split()[2]
    elif 'line vty' in x:
        mode = True
    elif mode and re.match(' ', x):
        if 'password' in x:
            ws['C10'] = x.split()[1]

# 「モジュール情報」の書き込み
row = 16
for x in module_data:
    ws.cell(row, 2).value = x['pid']
    ws.cell(row, 3).value = x['sn']
    ws.cell(row, 4).value = x['descr']
    row += 1

# 「インタフェース情報」の書き込み
row = 24
for x in interface_data:
    ws.cell(row, 2).value = x['interface']
    ws.cell(row, 3).value = x['link_status']
    
    if x['ip_address']:
        address = IPNetwork(x['ip_address'])
        ws.cell(row, 4).value = str(address.ip)
        ws.cell(row, 5).value = str(address.netmask)
    else:
        ws.cell(row, 4).value = '-'
        ws.cell(row, 5).value = '-'
    
    row += 1

# 「スタティックルート設定情報」の書き込み
row = 36
for x in config_static_data:
    if 'ip route' in x:
        ws.cell(row, 2).value = x.split()[2]
        ws.cell(row, 3).value = x.split()[3]
        ws.cell(row, 4).value = x.split()[4]
        
        row += 1

# 「ルーティングテーブル情報」の書き込み
row = 49
for x in route_data:
    ws.cell(row, 2).value = x['protocol']
    ws.cell(row, 3).value = x['network']
    
    mask = IPNetwork('0.0.0.0/' + x['mask'])
    
    ws.cell(row, 4).value = str(mask.netmask)
    ws.cell(row, 5).value = x['nexthop_ip'] or '-'
    ws.cell(row, 6).value = x['nexthop_if'] or '-'
    
    row += 1

wb.save(filename)

print('エクセルへの書き込みが完了しました')