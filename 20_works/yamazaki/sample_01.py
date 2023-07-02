from openpyxl import load_workbook
from netmiko import ConnectHandler
from textfsm import TextFSM
from netaddr import IPAddress, IPNetwork
import csv

# ホスト情報を読み込む
with open('hosts.csv', 'r') as f:
    reader = csv.DictReader(f)
    hosts = list(reader)

# 各ホストに接続し、情報を取得する
for host in hosts:
    
    # Excelワークブックを開く
    filename = host['filename']
    wb = load_workbook(filename)
    ws = wb['Tokyo']
    
    # 接続情報生成
    device = {
    'device_type': 'cisco_ios_telnet',
    'host': host['ip'],
    'password': host['password'],
    'secret': host['secret'],
    }
    
    # ホストへ接続
    connection = ConnectHandler(**device)
    
    # enableモードへ移行
    connection.enable()
    
##### 以下コマンド単位にモジュール化 #####
    
    # コマンド実行結果取得
    cmd = 'show interfaces'
    output = connection.send_command(cmd, use_textfsm = True)

    # 取得したデータをExcelに出力する
    row = 17
    for x in output:
        emp = '-'
        ws.cell(row, 2).value = x['interface']
        ws.cell(row, 3).value = x['link_status']
        
        if x['ip_address']:
            address = IPNetwork(x['ip_address'])
            ws.cell(row, 4).value = str(address.ip)
            ws.cell(row, 5).value = str(address.netmask)
        else:
            ws.cell(row, 4).value = emp
            ws.cell(row, 5).value = emp
        
        row +=1    
    
    # エクセルファイルの保存
    wb.save(filename)
    
    # ホスト接続の解除
    connection.disconnect()
    
##### コマンド単位にモジュール化 以上 #####
